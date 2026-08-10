from __future__ import annotations

from io import BytesIO
from pathlib import Path
from threading import Barrier, Event, Lock, Thread
from time import monotonic

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from pydantic import ValidationError
from sqlalchemy import select

from app.core.model_gateway import ModelGatewayError
from app.core.review_runner import (
    ReviewCellExecutionError,
    ReviewRunConfigurationError,
    ReviewRunner,
    _used_citations,
)
from app.models.review import ReviewCitation
from app.core.review_store import ReviewConflict, ReviewLimitExceeded, ReviewStore
from app.core.usage_budget import (
    UsageBudgetUnavailable,
    new_accounting_id,
)
from app.core.usage_budget_runtime import (
    ProviderUsageAttribution,
    TenantUsageBudgetOrchestrator,
)
from app.db.orm import TenantUsagePermitRow
from app.main import app
from app.models.review import (
    MAX_REVIEW_ANSWER_CHARS,
    MAX_REVIEW_CELLS,
    MAX_REVIEW_DOCUMENTS,
    MAX_REVIEW_QUESTION_CHARS,
    ReviewColumn,
    ReviewMatrix,
    ReviewMatrixCreateRequest,
    review_now,
)
from app.models.schemas import KnowledgeChunk, ModelConfig
from app.repositories.deps import get_store
from app.repositories.review_deps import get_review_runner, get_review_store


def headers(user_id: str = "user-jane") -> dict[str, str]:
    return {"x-aperture-user": user_id}


class ControlledGateway:
    def __init__(self, *, block: bool = False, answer: str | None = None) -> None:
        self.release = Event()
        if not block:
            self.release.set()
        self.three_active = Event()
        self.started = Event()
        self._lock = Lock()
        self.active = 0
        self.max_active = 0
        self.calls = 0
        self.messages: list[list[dict]] = []
        self.upstream_models: list[str] = []
        self.answer = answer

    def complete(self, *, route, messages, **_kwargs):
        with self._lock:
            self.active += 1
            self.calls += 1
            self.max_active = max(self.max_active, self.active)
            self.messages.append(messages)
            self.upstream_models.append(route.upstream_model)
            self.started.set()
            if self.active >= 3:
                self.three_active.set()
        try:
            if not self.release.wait(timeout=5):
                raise ModelGatewayError("test gateway timed out")
            user_prompt = str(messages[-1]["content"])
            if user_prompt.startswith("Question: reviewneedle forced failure"):
                raise ModelGatewayError("forced provider failure")
            answer = self.answer if self.answer is not None else "Supported answer [K1]."
            return {
                "choices": [{"message": {"role": "assistant", "content": answer}}],
                "usage": {"prompt_tokens": 7, "completion_tokens": 3, "total_tokens": 10},
            }
        finally:
            with self._lock:
                self.active -= 1


@pytest.fixture
def review_api():
    get_store.cache_clear()
    seed_store = get_store()
    _activate_model_provider(seed_store)
    _add_page_chunks(seed_store)
    review_store = ReviewStore(":memory:")
    gateway = ControlledGateway(block=True)
    runner = ReviewRunner(
        review_store,
        seed_store,
        max_concurrency=3,
        gateway_factory=lambda: gateway,
        activity_writes_sql_only=True,
    )
    app.dependency_overrides[get_review_store] = lambda: review_store
    app.dependency_overrides[get_review_runner] = lambda: runner
    api = TestClient(app)
    yield api, seed_store, review_store, runner, gateway
    gateway.release.set()
    runner.wait_for_idle(timeout=5)
    runner.close()
    review_store.close()
    app.dependency_overrides.pop(get_review_store, None)
    app.dependency_overrides.pop(get_review_runner, None)
    get_store.cache_clear()


def test_three_by_three_run_has_terminal_answers_errors_pages_and_bounded_concurrency(
    review_api,
) -> None:
    api, seed_store, _review_store, runner, gateway = review_api
    created = api.post("/api/review/matrices", json=_matrix_payload(), headers=headers())

    assert created.status_code == 201
    matrix_id = created.json()["id"]
    accepted = api.post(f"/api/review/matrices/{matrix_id}/run", headers=headers())
    assert accepted.status_code == 202
    assert accepted.json() == {"matrix_id": matrix_id, "status": "running", "cell_count": 9}
    assert gateway.three_active.wait(timeout=2)

    duplicate = api.post(f"/api/review/matrices/{matrix_id}/run", headers=headers())
    mutation = api.patch(
        f"/api/review/matrices/{matrix_id}",
        json={"name": "Changed while running"},
        headers=headers(),
    )
    deletion = api.delete(f"/api/review/matrices/{matrix_id}", headers=headers())

    assert duplicate.status_code == 409
    assert mutation.status_code == 409
    assert deletion.status_code == 409

    gateway.release.set()
    runner.wait_for_idle(timeout=5)
    detail = api.get(f"/api/review/matrices/{matrix_id}", headers=headers())

    assert detail.status_code == 200
    body = detail.json()
    assert body["matrix"]["status"] == "complete"
    assert len(body["cells"]) == 9
    assert {cell["status"] for cell in body["cells"]} == {"complete", "failed"}
    assert all(cell["answer"] or cell["error"] for cell in body["cells"])
    complete_cells = [cell for cell in body["cells"] if cell["status"] == "complete"]
    assert complete_cells
    assert all("run_token" not in cell for cell in body["cells"])
    assert all(cell["citations"][0]["page_start"] in {2, 3, 4} for cell in complete_cells)
    assert all(cell["citations"][0]["locator"].startswith("Paragraph") for cell in complete_cells)
    assert gateway.max_active == 3
    assert set(gateway.upstream_models) == {"gpt-4o-mini"}
    assert all("[K1] (p." in str(messages[-1]["content"]) for messages in gateway.messages)

    review_usage = [record for record in seed_store.usage_records if record.surface == "review"]
    assert len(review_usage) == 6
    assert all(record.total_tokens == 10 for record in review_usage)
    review_audits = [event for event in seed_store.audit_events if event.action == "review.run"]
    assert len(review_audits) == 1
    assert review_audits[0].target == matrix_id

    csv_response = api.get(f"/api/review/matrices/{matrix_id}/export.csv", headers=headers())
    xlsx_response = api.get(f"/api/review/matrices/{matrix_id}/export.xlsx", headers=headers())
    assert csv_response.status_code == 200
    assert "Document,Deadline,Exposure,Failure check,Sources" in csv_response.text
    assert "p. " in csv_response.text
    assert "Deadline: [K1]" in csv_response.text
    assert "Exposure: [K1]" in csv_response.text
    assert xlsx_response.status_code == 200
    workbook = load_workbook(BytesIO(xlsx_response.content), read_only=True)
    values = list(workbook["Review"].values)
    assert values[0] == ("Document", "Deadline", "Exposure", "Failure check", "Sources")
    assert len(values) == 4


def test_review_routes_enforce_source_acl_owner_scope_and_null_matter(review_api) -> None:
    api, *_rest = review_api
    inaccessible = _matrix_payload()
    inaccessible["knowledge_config_ids"] = ["knowledge-policy-library"]
    inaccessible["document_ids"] = ["doc-drive-ai-use-policy"]

    blocked = api.post("/api/review/matrices", json=inaccessible, headers=headers())
    assert blocked.status_code == 403

    created = api.post("/api/review/matrices", json=_matrix_payload(), headers=headers())
    matrix_id = created.json()["id"]
    cross_user = api.get(f"/api/review/matrices/{matrix_id}", headers=headers("user-casey"))
    assert cross_user.status_code == 404

    with_matter = _matrix_payload()
    with_matter["matter_id"] = "matter-not-yet-supported"
    rejected_matter = api.post("/api/review/matrices", json=with_matter, headers=headers())
    assert rejected_matter.status_code == 422


def test_document_acl_revocation_blocks_read_export_but_not_owner_cleanup(review_api) -> None:
    api, seed_store, *_rest = review_api
    created = api.post(
        "/api/review/matrices", json=_matrix_payload(only_one=True), headers=headers()
    )
    matrix_id = created.json()["id"]
    document = next(
        document
        for document in seed_store.knowledge_documents_for("knowledge-box-matters")
        if document.id == "doc-box-complaint-outline"
    )
    seed_store.vector_store.upsert_sources(
        [document.model_copy(update={"acl_group_ids": ["group-corporate"]})],
        [],
    )

    assert api.get(f"/api/review/matrices/{matrix_id}", headers=headers()).status_code == 403
    assert (
        api.get(f"/api/review/matrices/{matrix_id}/export.csv", headers=headers()).status_code
        == 403
    )
    assert api.get("/api/review/matrices", headers=headers()).json() == []
    assert api.delete(f"/api/review/matrices/{matrix_id}", headers=headers()).status_code == 200


def test_tenant_admin_document_acl_matches_vector_visibility(review_api) -> None:
    api, seed_store, *_rest = review_api
    document = next(
        document
        for document in seed_store.knowledge_documents_for("knowledge-box-matters")
        if document.id == "doc-box-complaint-outline"
    )
    seed_store.vector_store.upsert_sources(
        [document.model_copy(update={"acl_group_ids": ["group-corporate"]})],
        [],
    )
    payload = _matrix_payload(only_one=True)

    response = api.post("/api/review/matrices", json=payload, headers=headers("user-admin"))

    assert response.status_code == 403
    assert response.json()["detail"] == "Review document access is restricted by source ACL policy."


def test_review_content_filter_configuration_fails_closed_before_dispatch(review_api) -> None:
    api, seed_store, _review_store, _runner, gateway = review_api
    seed_store.models["gpt-4o-mini"].content_filter_ids = ["missing-filter"]
    created = api.post(
        "/api/review/matrices", json=_matrix_payload(only_one=True), headers=headers()
    )
    matrix_id = created.json()["id"]

    response = api.post(f"/api/review/matrices/{matrix_id}/run", headers=headers())

    assert response.status_code == 503
    assert "content filters are unavailable" in response.json()["detail"]
    assert gateway.calls == 0
    detail = api.get(f"/api/review/matrices/{matrix_id}", headers=headers()).json()
    assert detail["matrix"]["status"] == "draft"
    assert detail["cells"] == []


def test_review_unknown_openrouter_alias_never_calls_provider(
    review_api,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    api, seed_store, review_store, _runner, gateway = review_api
    model_id = "review-alias-without-upstream"
    seed_store.models[model_id] = ModelConfig(
        id=model_id,
        provider_id="provider-openrouter",
        provider_name="OpenRouter",
        name="Review alias without upstream",
        upstream_model_id=None,
        platform_enabled=True,
        group_ids=["group-litigation"],
    )

    def unexpected_credential_read(*_args: object, **_kwargs: object) -> None:
        raise AssertionError("unknown selected model must fail before credential resolution")

    monkeypatch.setattr(
        seed_store,
        "provider_key_secret_for_provider",
        unexpected_credential_read,
    )
    payload = _matrix_payload(only_one=True)
    payload["model_id"] = model_id
    created = api.post("/api/review/matrices", json=payload, headers=headers())
    assert created.status_code == 201

    matrix_id = created.json()["id"]
    response = api.post(f"/api/review/matrices/{matrix_id}/run", headers=headers())

    assert response.status_code == 503
    assert response.json()["detail"] == (
        f"Selected model '{model_id}' has no explicit OpenRouter upstream model id."
    )
    assert seed_store.openrouter_default_model not in response.text
    assert gateway.calls == 0
    assert review_store.get_matrix(matrix_id).status == "draft"
    assert review_store.cells_for_matrix(matrix_id) == []


def test_review_run_and_rerun_gate_require_sql_activity_sink_capability(review_api) -> None:
    api, seed_store, review_store, runner, gateway = review_api
    blocked_runner = ReviewRunner(
        review_store,
        seed_store,
        gateway_factory=lambda: gateway,
    )
    app.dependency_overrides[get_review_runner] = lambda: blocked_runner
    created = api.post(
        "/api/review/matrices",
        json=_matrix_payload(only_one=True),
        headers=headers(),
    )
    matrix_id = created.json()["id"]

    blocked_run = api.post(f"/api/review/matrices/{matrix_id}/run", headers=headers())

    assert blocked_run.status_code == 503
    assert "SQL audit, usage, and export-outbox sink" in blocked_run.json()["detail"]
    assert review_store.get_matrix(matrix_id).status == "draft"
    assert review_store.cells_for_matrix(matrix_id) == []
    assert gateway.calls == 0

    # Exercise rerun through a terminal cell prepared by the explicitly enabled
    # test runner, then swap the gated runner back in before the request.
    app.dependency_overrides[get_review_runner] = lambda: runner
    gateway.release.set()
    assert api.post(f"/api/review/matrices/{matrix_id}/run", headers=headers()).status_code == 202
    runner.wait_for_idle(timeout=5)
    cell_id = review_store.cells_for_matrix(matrix_id)[0].id
    app.dependency_overrides[get_review_runner] = lambda: blocked_runner
    blocked_rerun = api.post(
        f"/api/review/matrices/{matrix_id}/cells/{cell_id}/rerun",
        headers=headers(),
    )
    assert blocked_rerun.status_code == 503
    assert review_store.get_cell(matrix_id, cell_id).attempt == 1
    blocked_runner.close()
    app.dependency_overrides[get_review_runner] = lambda: runner


def test_terminal_cell_can_be_rerun_and_tracks_a_new_attempt(review_api) -> None:
    api, seed_store, _review_store, runner, gateway = review_api
    gateway.release.set()
    created = api.post(
        "/api/review/matrices",
        json=_matrix_payload(only_one=True),
        headers=headers(),
    )
    matrix_id = created.json()["id"]
    first_run = api.post(f"/api/review/matrices/{matrix_id}/run", headers=headers())
    assert first_run.status_code == 202
    runner.wait_for_idle(timeout=5)
    first_cell = api.get(f"/api/review/matrices/{matrix_id}", headers=headers()).json()["cells"][0]
    assert first_cell["status"] == "complete"
    assert first_cell["attempt"] == 1

    rerun = api.post(
        f"/api/review/matrices/{matrix_id}/cells/{first_cell['id']}/rerun",
        headers=headers(),
    )
    assert rerun.status_code == 202
    assert rerun.json()["cell_count"] == 1
    runner.wait_for_idle(timeout=5)

    detail = api.get(f"/api/review/matrices/{matrix_id}", headers=headers()).json()
    assert detail["matrix"]["status"] == "complete"
    assert detail["cells"][0]["status"] == "complete"
    assert detail["cells"][0]["attempt"] == 2
    review_usage = [record for record in seed_store.usage_records if record.surface == "review"]
    assert len(review_usage) == 2
    review_audits = [event for event in seed_store.audit_events if event.action == "review.run"]
    assert [event.metadata["mode"] for event in review_audits] == ["matrix", "cell_rerun"]


def test_review_exports_escape_formula_like_column_labels(review_api) -> None:
    api, _seed_store, _review_store, runner, gateway = review_api
    gateway.release.set()
    payload = _matrix_payload(only_one=True)
    payload["columns"][0]["label"] = "=WEBSERVICE(bad)"
    created = api.post("/api/review/matrices", json=payload, headers=headers())
    matrix_id = created.json()["id"]
    assert api.post(f"/api/review/matrices/{matrix_id}/run", headers=headers()).status_code == 202
    runner.wait_for_idle(timeout=5)

    csv_response = api.get(f"/api/review/matrices/{matrix_id}/export.csv", headers=headers())
    xlsx_response = api.get(f"/api/review/matrices/{matrix_id}/export.xlsx", headers=headers())

    assert "'=WEBSERVICE(bad)" in csv_response.text
    workbook = load_workbook(BytesIO(xlsx_response.content), read_only=True)
    assert workbook["Review"]["B1"].value == "'=WEBSERVICE(bad)"


def test_review_answer_size_cap_fails_cell_honestly() -> None:
    seed_store = _fresh_seed_store()
    review_store = ReviewStore(":memory:")
    gateway = ControlledGateway(answer="x" * (MAX_REVIEW_ANSWER_CHARS + 1))
    runner = ReviewRunner(
        review_store,
        seed_store,
        gateway_factory=lambda: gateway,
        activity_writes_sql_only=True,
    )
    matrix = _direct_matrix()
    review_store.create_matrix(matrix)

    runner.start_matrix(matrix.id, seed_store.users["user-jane"])
    runner.wait_for_idle(timeout=5)

    cell = review_store.cells_for_matrix(matrix.id)[0]
    assert cell.status == "failed"
    assert cell.answer is None
    assert "character review cell limit" in str(cell.error)
    runner.close()
    review_store.close()
    get_store.cache_clear()


@pytest.mark.parametrize(
    ("answer", "expected_error"),
    [
        ("An uncited answer.", "did not include a verifiable"),
        ("An invented source [K999].", "was not supplied"),
    ],
)
def test_review_requires_only_verifiable_inline_citations(answer: str, expected_error: str) -> None:
    seed_store = _fresh_seed_store()
    review_store = ReviewStore(":memory:")
    gateway = ControlledGateway(answer=answer)
    runner = ReviewRunner(
        review_store,
        seed_store,
        gateway_factory=lambda: gateway,
        activity_writes_sql_only=True,
    )
    matrix = _direct_matrix()
    review_store.create_matrix(matrix)

    runner.start_matrix(matrix.id, seed_store.users["user-jane"])
    runner.wait_for_idle(timeout=5)

    cell = review_store.cells_for_matrix(matrix.id)[0]
    assert cell.status == "failed"
    assert expected_error in str(cell.error)
    runner.close()
    review_store.close()
    get_store.cache_clear()


def test_gateway_factory_failure_does_not_reserve_a_run() -> None:
    seed_store = _fresh_seed_store()
    review_store = ReviewStore(":memory:")

    def broken_gateway():
        raise RuntimeError("factory unavailable")

    runner = ReviewRunner(
        review_store,
        seed_store,
        gateway_factory=broken_gateway,
        activity_writes_sql_only=True,
    )
    matrix = _direct_matrix()
    review_store.create_matrix(matrix)

    with pytest.raises(ReviewRunConfigurationError, match="no work was accepted"):
        runner.start_matrix(matrix.id, seed_store.users["user-jane"])

    assert review_store.get_matrix(matrix.id).status == "draft"
    assert review_store.cells_for_matrix(matrix.id) == []
    runner.close()
    review_store.close()
    get_store.cache_clear()


def test_platform_owner_review_usage_and_audit_are_attributed_to_matrix_tenant() -> None:
    seed_store = _fresh_seed_store()
    review_store = ReviewStore(":memory:")
    gateway = ControlledGateway()
    runner = ReviewRunner(
        review_store,
        seed_store,
        gateway_factory=lambda: gateway,
        activity_writes_sql_only=True,
    )
    matrix = _direct_matrix().model_copy(update={"owner_user_id": "user-owner"})
    review_store.create_matrix(matrix)

    runner.start_matrix(matrix.id, seed_store.users["user-owner"])
    runner.wait_for_idle(timeout=5)

    assert seed_store.usage_records[-1].tenant_id == matrix.tenant_id
    review_audit = next(
        event for event in reversed(seed_store.audit_events) if event.action == "review.run"
    )
    assert review_audit.tenant_id == matrix.tenant_id
    assert review_audit.actor_id == "user-owner"
    runner.close()
    review_store.close()
    get_store.cache_clear()


def test_review_runner_rejects_ordinary_cross_tenant_actor_before_reservation() -> None:
    seed_store = _fresh_seed_store()
    review_store = ReviewStore(":memory:")
    gateway = ControlledGateway()
    runner = ReviewRunner(
        review_store,
        seed_store,
        gateway_factory=lambda: gateway,
        activity_writes_sql_only=True,
    )
    matrix = _direct_matrix(matrix_id="review-cross-tenant")
    review_store.create_matrix(matrix)
    cross_tenant_actor = seed_store.users["user-jane"].model_copy(
        update={"tenant_id": "tenant-other"}
    )

    with pytest.raises(ReviewRunConfigurationError, match="not in the review matrix tenant"):
        runner.start_matrix(matrix.id, cross_tenant_actor)

    assert review_store.get_matrix(matrix.id).status == "draft"
    assert review_store.cells_for_matrix(matrix.id) == []
    assert gateway.calls == 0
    with seed_store.application_state_repository.engine.connect() as connection:
        statuses = connection.execute(select(TenantUsagePermitRow.status)).scalars().all()
    assert statuses == []
    runner.close()
    review_store.close()
    get_store.cache_clear()


def test_review_budget_exhaustion_returns_429_before_provider_io(review_api) -> None:
    api, seed_store, _review_store, _runner, gateway = review_api
    actor = seed_store.users["user-jane"]
    repository = seed_store.usage_budget_repository
    repository.set_budget("tenant-example", 1, updated_by=actor.id)
    usage_context = TenantUsageBudgetOrchestrator(repository).begin_request(
        actor=actor,
        request_id=new_accounting_id(),
        resource_tenant_id="tenant-example",
        known_tenant_ids=seed_store.tenants.keys(),
    )
    usage_context.settle_provider_child(
        completion_id=new_accounting_id(),
        usage={"total_tokens": 1},
        attribution=ProviderUsageAttribution(
            model_id="gpt-4o-mini",
            provider_name="Test Provider",
            surface="review-budget-setup",
        ),
    )
    usage_context.complete_success()

    created = api.post(
        "/api/review/matrices",
        json=_matrix_payload(only_one=True),
        headers=headers(),
    )
    response = api.post(
        f"/api/review/matrices/{created.json()['id']}/run",
        headers=headers(),
    )

    assert response.status_code == 429
    assert int(response.headers["retry-after"]) > 0
    assert gateway.calls == 0


def test_review_lazy_admission_blocks_queued_cells_after_first_crosses_budget() -> None:
    seed_store = _fresh_seed_store()
    review_store = ReviewStore(":memory:")
    gateway = ControlledGateway()
    actor = seed_store.users["user-jane"]
    seed_store.usage_budget_repository.set_budget(
        "tenant-example",
        1,
        updated_by=actor.id,
    )
    runner = ReviewRunner(
        review_store,
        seed_store,
        max_concurrency=1,
        gateway_factory=lambda: gateway,
        activity_writes_sql_only=True,
    )
    matrix = _direct_matrix(matrix_id="review-lazy-admission").model_copy(
        update={
            "columns": [
                ReviewColumn(
                    id="deadline",
                    label="Deadline",
                    question="reviewneedle response deadline",
                ),
                ReviewColumn(
                    id="exposure",
                    label="Exposure",
                    question="reviewneedle exposure amount",
                ),
            ]
        }
    )
    review_store.create_matrix(matrix)

    runner.start_matrix(matrix.id, actor)
    assert runner.wait_for_idle(timeout=5)

    cells = {cell.column_id: cell for cell in review_store.cells_for_matrix(matrix.id)}
    assert gateway.calls == 1
    assert cells["deadline"].status == "complete"
    assert cells["exposure"].status == "failed"
    assert "daily token budget" in str(cells["exposure"].error).lower()
    assert review_store.get_matrix(matrix.id).status == "complete"
    records = [record for record in seed_store.usage_records if record.surface == "review"]
    assert len(records) == 1
    assert records[0].total_tokens == 10
    with seed_store.application_state_repository.engine.connect() as connection:
        statuses = connection.execute(select(TenantUsagePermitRow.status)).scalars().all()
    assert statuses == ["completed"]
    runner.close()
    review_store.close()
    get_store.cache_clear()


def test_review_permit_close_failure_never_delivers_a_complete_cell() -> None:
    seed_store = _fresh_seed_store()
    review_store = ReviewStore(":memory:")
    gateway = ControlledGateway()

    class CloseFailureRepository:
        def __init__(self):
            self.delegate = seed_store.usage_budget_repository

        def __getattr__(self, name):
            return getattr(self.delegate, name)

        def complete_permit(self, permit_id: str, *, now=None):
            raise UsageBudgetUnavailable("forced permit-close failure")

    runner = ReviewRunner(
        review_store,
        seed_store,
        gateway_factory=lambda: gateway,
        activity_writes_sql_only=True,
        usage_budget_orchestrator=TenantUsageBudgetOrchestrator(CloseFailureRepository()),
    )
    matrix = _direct_matrix(matrix_id="review-close-failure")
    review_store.create_matrix(matrix)

    runner.start_matrix(matrix.id, seed_store.users["user-jane"])
    assert runner.wait_for_idle(timeout=5)

    cell = review_store.cells_for_matrix(matrix.id)[0]
    assert cell.status == "failed"
    assert cell.answer is None
    assert "usage accounting is unavailable" in str(cell.error).lower()
    assert len(seed_store.usage_records) == 1
    assert seed_store.usage_records[0].total_tokens == 10
    with seed_store.application_state_repository.engine.connect() as connection:
        statuses = connection.execute(select(TenantUsagePermitRow.status)).scalars().all()
    assert statuses == ["failed"]
    runner.close()
    review_store.close()
    get_store.cache_clear()


@pytest.mark.parametrize("after_transition", [False, True], ids=["pending", "running"])
def test_unexpected_review_claim_failure_is_terminal_before_provider_io(
    monkeypatch,
    after_transition: bool,
) -> None:
    seed_store = _fresh_seed_store()
    review_store = ReviewStore(":memory:")
    gateway = ControlledGateway()
    runner = ReviewRunner(
        review_store,
        seed_store,
        gateway_factory=lambda: gateway,
        activity_writes_sql_only=True,
    )
    matrix = _direct_matrix(matrix_id="review-claim-failure")
    review_store.create_matrix(matrix)

    original_mark_cell_running = review_store.mark_cell_running

    def fail_claim(*args, **kwargs):
        if after_transition:
            original_mark_cell_running(*args, **kwargs)
        raise RuntimeError("forced claim failure")

    monkeypatch.setattr(review_store, "mark_cell_running", fail_claim)

    runner.start_matrix(matrix.id, seed_store.users["user-jane"])
    assert runner.wait_for_idle(timeout=5)

    assert gateway.calls == 0
    assert list(seed_store.usage_records) == []
    cell = review_store.cells_for_matrix(matrix.id)[0]
    assert cell.status == "failed"
    assert "RuntimeError" in str(cell.error)
    assert review_store.get_matrix(matrix.id).status == "complete"
    with seed_store.application_state_repository.engine.connect() as connection:
        statuses = connection.execute(select(TenantUsagePermitRow.status)).scalars().all()
    assert statuses == ["failed"]
    runner.close()
    review_store.close()
    get_store.cache_clear()


def test_bounded_runner_shutdown_marks_blocked_work_interrupted() -> None:
    seed_store = _fresh_seed_store()
    review_store = ReviewStore(":memory:")
    gateway = ControlledGateway(block=True)
    runner = ReviewRunner(
        review_store,
        seed_store,
        gateway_factory=lambda: gateway,
        activity_writes_sql_only=True,
    )
    matrix = _direct_matrix()
    review_store.create_matrix(matrix)
    runner.start_matrix(matrix.id, seed_store.users["user-jane"])
    assert gateway.started.wait(timeout=2)

    started = monotonic()
    runner.close(timeout=0.01)
    elapsed = monotonic() - started

    assert elapsed < 0.5
    assert review_store.get_matrix(matrix.id).status == "interrupted"
    assert review_store.cells_for_matrix(matrix.id)[0].status == "interrupted"
    gateway.release.set()
    runner.wait_for_idle(timeout=5)
    assert review_store.get_matrix(matrix.id).status == "interrupted"
    records = [record for record in seed_store.usage_records if record.surface == "review"]
    assert len(records) == 1
    assert records[0].total_tokens == 10
    with seed_store.application_state_repository.engine.connect() as connection:
        statuses = connection.execute(select(TenantUsagePermitRow.status)).scalars().all()
    assert statuses == ["abandoned"]
    review_store.close()
    get_store.cache_clear()


def test_submit_and_close_register_future_atomically(monkeypatch) -> None:
    seed_store = _fresh_seed_store()
    review_store = ReviewStore(":memory:")
    gateway = ControlledGateway(block=True)
    runner = ReviewRunner(
        review_store,
        seed_store,
        gateway_factory=lambda: gateway,
        activity_writes_sql_only=True,
    )
    matrix = _direct_matrix(matrix_id="review-submit-close-race")
    review_store.create_matrix(matrix)

    claim_entered = Event()
    release_claim = Event()
    submit_entered = Event()
    release_submit = Event()
    close_done = Event()
    original_claim = runner._claim_cell
    original_submit = runner._executor.submit

    def blocked_claim(cell):
        claim_entered.set()
        assert release_claim.wait(timeout=5)
        return original_claim(cell)

    def blocked_submit(*args, **kwargs):
        future = original_submit(*args, **kwargs)
        submit_entered.set()
        assert release_submit.wait(timeout=5)
        return future

    monkeypatch.setattr(runner, "_claim_cell", blocked_claim)
    monkeypatch.setattr(runner._executor, "submit", blocked_submit)
    start_errors: list[Exception] = []
    close_errors: list[Exception] = []
    close_elapsed: list[float] = []

    def start_run() -> None:
        try:
            runner.start_matrix(matrix.id, seed_store.users["user-jane"])
        except Exception as exc:  # noqa: BLE001 - surfaced below after thread cleanup
            start_errors.append(exc)

    def close_runner() -> None:
        started = monotonic()
        try:
            runner.close(timeout=0.01)
        except Exception as exc:  # noqa: BLE001 - surfaced below after thread cleanup
            close_errors.append(exc)
        finally:
            close_elapsed.append(monotonic() - started)
            close_done.set()

    start_thread = Thread(target=start_run)
    close_thread = Thread(target=close_runner)
    start_thread.start()
    assert submit_entered.wait(timeout=2)
    assert claim_entered.wait(timeout=2)
    close_thread.start()
    try:
        assert not close_done.wait(timeout=0.05)
        release_submit.set()
        assert close_done.wait(timeout=1)
    finally:
        release_submit.set()
        release_claim.set()
        gateway.release.set()
        start_thread.join(timeout=2)
        close_thread.join(timeout=2)

    assert start_errors == []
    assert close_errors == []
    assert close_elapsed[0] < 0.5
    assert runner.wait_for_idle(timeout=5)
    assert gateway.calls == 0
    assert review_store.get_matrix(matrix.id).status == "interrupted"
    assert review_store.cells_for_matrix(matrix.id)[0].status == "interrupted"
    with seed_store.application_state_repository.engine.connect() as connection:
        statuses = connection.execute(select(TenantUsagePermitRow.status)).scalars().all()
    assert statuses == ["abandoned"]
    review_store.close()
    get_store.cache_clear()


def test_tenant_and_owner_purge_cancel_workers_and_prevent_late_writes() -> None:
    seed_store = _fresh_seed_store()
    review_store = ReviewStore(":memory:")
    gateway = ControlledGateway(block=True)
    runner = ReviewRunner(
        review_store,
        seed_store,
        gateway_factory=lambda: gateway,
        activity_writes_sql_only=True,
    )
    running = _direct_matrix(matrix_id="purge-running")
    same_tenant = _direct_matrix(matrix_id="purge-same-tenant").model_copy(
        update={"owner_user_id": "user-casey"}
    )
    other_tenant = _direct_matrix(matrix_id="purge-other-tenant").model_copy(
        update={"tenant_id": "tenant-other"}
    )
    for matrix in (running, same_tenant, other_tenant):
        review_store.create_matrix(matrix)
    runner.start_matrix(running.id, seed_store.users["user-jane"])
    assert gateway.started.wait(timeout=2)
    usage_before = len(seed_store.usage_records)

    assert runner.purge_tenant("tenant-example") == 2
    assert review_store.get_matrix(running.id) is None
    assert review_store.get_matrix(same_tenant.id) is None
    assert review_store.get_matrix(other_tenant.id) is not None
    gateway.release.set()
    assert runner.wait_for_idle(timeout=5)

    assert review_store.get_matrix(running.id) is None
    assert review_store.cells_for_matrix(running.id) == []
    # The raw provider call succeeded before cancellation was observed, so its
    # exact usage must survive even though the purged cell cannot accept output.
    assert len(seed_store.usage_records) == usage_before + 1
    with seed_store.application_state_repository.engine.connect() as connection:
        statuses = connection.execute(select(TenantUsagePermitRow.status)).scalars().all()
    assert statuses == ["abandoned"]
    assert runner.purge_owner("user-jane") == 1
    assert review_store.get_matrix(other_tenant.id) is None
    assert runner.purge_owner("user-jane") == 0
    runner.close()
    review_store.close()
    get_store.cache_clear()


def test_main_lifespan_opens_review_then_closes_it_before_seed_store(monkeypatch) -> None:
    import app.main as main_module

    events: list[str] = []

    class FakeUsageBudgetRepository:
        def abandon_started_permits(self) -> int:
            events.append("usage-reconcile")
            return 0

    class FakeSeedStore:
        usage_budget_repository = FakeUsageBudgetRepository()

        def resume_identity_cleanup_jobs(self) -> int:
            events.append("identity-cleanup-resume")
            return 0

        def close(self) -> None:
            events.append("seed-close")

    monkeypatch.setattr(main_module, "get_store", lambda: FakeSeedStore())
    monkeypatch.setattr(
        main_module,
        "initialize_review_services",
        lambda: events.append("review-open"),
    )
    monkeypatch.setattr(
        main_module,
        "close_review_services",
        lambda: events.append("review-close"),
    )

    with TestClient(main_module.app):
        assert events == ["usage-reconcile", "review-open", "identity-cleanup-resume"]

    assert events == [
        "usage-reconcile",
        "review-open",
        "identity-cleanup-resume",
        "review-close",
        "seed-close",
    ]


def test_review_store_persists_and_interrupts_unfinished_run(tmp_path: Path) -> None:
    path = tmp_path / "reviews.sqlite3"
    first = ReviewStore(str(path))
    matrix = _direct_matrix()
    first.create_matrix(matrix)
    running, cells = first.begin_run(matrix.id)
    assert running.status == "running"
    first.mark_cell_running(
        matrix.id,
        cells[0].id,
        expected_attempt=cells[0].attempt,
        run_token=cells[0].run_token,
    )
    first.close()

    restarted = ReviewStore(str(path))
    persisted = restarted.get_matrix(matrix.id)
    assert persisted is not None
    assert persisted.status == "interrupted"
    interrupted_cells = restarted.cells_for_matrix(matrix.id)
    assert interrupted_cells
    assert {cell.status for cell in interrupted_cells} == {"interrupted"}
    assert all(cell.error and "interrupted" in cell.error.lower() for cell in interrupted_cells)
    restarted.close()


def test_attempt_token_blocks_stale_worker_after_restart_and_rerun(tmp_path: Path) -> None:
    path = tmp_path / "attempt-cas.sqlite3"
    old_process = ReviewStore(str(path))
    matrix = _direct_matrix()
    old_process.create_matrix(matrix)
    _running, old_cells = old_process.begin_run(matrix.id)
    old_cell = old_cells[0]
    old_process.mark_cell_running(
        matrix.id,
        old_cell.id,
        expected_attempt=old_cell.attempt,
        run_token=old_cell.run_token,
    )

    restarted = ReviewStore(str(path))
    interrupted = restarted.get_matrix(matrix.id)
    rerunning, new_cell = restarted.begin_cell_rerun(
        matrix.id,
        old_cell.id,
        expected_updated_at=interrupted.updated_at,
    )
    assert rerunning.status == "running"
    restarted.mark_cell_running(
        matrix.id,
        new_cell.id,
        expected_attempt=new_cell.attempt,
        run_token=new_cell.run_token,
    )

    with pytest.raises(ReviewConflict, match="stale"):
        old_process.complete_cell(
            matrix.id,
            old_cell.id,
            answer="Old result [K1]",
            citations=[],
            expected_attempt=old_cell.attempt,
            run_token=old_cell.run_token,
        )
    with pytest.raises(ReviewConflict, match="stale"):
        old_process.fail_reserved_cell_attempt(
            matrix.id,
            old_cell.id,
            error="Stale failure",
            expected_attempt=old_cell.attempt,
            run_token=old_cell.run_token,
        )

    current = restarted.get_cell(matrix.id, new_cell.id)
    assert current.status == "running"
    assert current.attempt == 2
    assert current.run_token == new_cell.run_token
    restarted.fail_cell(
        matrix.id,
        new_cell.id,
        error="test cleanup",
        expected_attempt=new_cell.attempt,
        run_token=new_cell.run_token,
    )
    restarted.finish_matrix_if_terminal(matrix.id)
    old_process.close()
    restarted.close()


def test_matrix_update_compare_and_set_rejects_stale_editor() -> None:
    review_store = ReviewStore(":memory:")
    matrix = _direct_matrix()
    review_store.create_matrix(matrix)
    first_read = review_store.get_matrix(matrix.id)
    second_read = review_store.get_matrix(matrix.id)
    first_update = first_read.model_copy(update={"name": "First", "updated_at": review_now()})
    second_update = second_read.model_copy(update={"name": "Second", "updated_at": review_now()})

    review_store.update_matrix(
        first_update,
        clear_cells=False,
        expected_updated_at=first_read.updated_at,
    )
    with pytest.raises(ReviewConflict, match="changed while"):
        review_store.update_matrix(
            second_update,
            clear_cells=False,
            expected_updated_at=second_read.updated_at,
        )
    assert review_store.get_matrix(matrix.id).name == "First"
    review_store.close()


def test_duplicate_run_reservation_is_atomic_across_sqlite_connections(tmp_path: Path) -> None:
    path = tmp_path / "duplicate-run.sqlite3"
    first = ReviewStore(str(path))
    second = ReviewStore(str(path))
    matrix = _direct_matrix()
    first.create_matrix(matrix)
    observed = first.get_matrix(matrix.id)
    barrier = Barrier(2)
    outcomes: list[str] = []
    outcome_lock = Lock()

    def reserve(review_store: ReviewStore) -> None:
        barrier.wait(timeout=2)
        try:
            review_store.begin_run(
                matrix.id,
                expected_updated_at=observed.updated_at,
            )
            outcome = "accepted"
        except ReviewConflict:
            outcome = "conflict"
        with outcome_lock:
            outcomes.append(outcome)

    threads = [Thread(target=reserve, args=(review_store,)) for review_store in (first, second)]
    for thread in threads:
        thread.start()
    for thread in threads:
        thread.join(timeout=3)

    assert sorted(outcomes) == ["accepted", "conflict"]
    first.interrupt_unfinished_runs()
    first.close()
    second.close()


def test_review_store_atomically_bounds_queue_and_rejects_active_mutations() -> None:
    review_store = ReviewStore(":memory:", max_queued_runs=1)
    first = _direct_matrix(matrix_id="matrix-one")
    second = _direct_matrix(matrix_id="matrix-two")
    review_store.create_matrix(first)
    review_store.create_matrix(second)
    review_store.begin_run(first.id)

    with pytest.raises(ReviewConflict, match="active run"):
        review_store.begin_run(first.id)
    with pytest.raises(ReviewLimitExceeded, match="limited to 1"):
        review_store.begin_run(second.id)
    with pytest.raises(ReviewConflict, match="cannot be changed"):
        review_store.update_matrix(first.model_copy(update={"name": "No"}), clear_cells=False)
    with pytest.raises(ReviewConflict, match="cannot be deleted"):
        review_store.delete_matrix(first.id)
    review_store.close()


def test_review_request_hard_limits() -> None:
    base = _matrix_payload()
    too_many_documents = {
        **base,
        "document_ids": [f"doc-{index}" for index in range(MAX_REVIEW_DOCUMENTS + 1)],
    }
    with pytest.raises(ValidationError):
        ReviewMatrixCreateRequest.model_validate(too_many_documents)

    too_long_question = _matrix_payload(only_one=True)
    too_long_question["columns"][0]["question"] = "q" * (MAX_REVIEW_QUESTION_CHARS + 1)
    with pytest.raises(ValidationError):
        ReviewMatrixCreateRequest.model_validate(too_long_question)

    too_many_cells = {
        **base,
        "document_ids": [f"doc-{index}" for index in range(20)],
        "columns": [
            {"id": f"column-{index}", "label": f"C{index}", "question": "Question"}
            for index in range((MAX_REVIEW_CELLS // 20) + 1)
        ],
    }
    with pytest.raises(ValidationError, match="limited"):
        ReviewMatrixCreateRequest.model_validate(too_many_cells)


def _matrix_payload(*, only_one: bool = False) -> dict:
    documents = [
        "doc-box-complaint-outline",
        "doc-box-discovery-plan",
        "doc-box-client-update",
    ]
    columns = [
        {"id": "deadline", "label": "Deadline", "question": "reviewneedle response deadline"},
        {"id": "exposure", "label": "Exposure", "question": "reviewneedle exposure amount"},
        {"id": "failure", "label": "Failure check", "question": "reviewneedle forced failure"},
    ]
    if only_one:
        documents = documents[:1]
        columns = columns[:1]
    return {
        "name": "Complaint Matrix",
        "knowledge_config_ids": ["knowledge-box-matters"],
        "document_ids": documents,
        "columns": columns,
        "model_id": "gpt-4o-mini",
        "matter_id": None,
    }


def _direct_matrix(*, matrix_id: str = "review-direct") -> ReviewMatrix:
    return ReviewMatrix(
        id=matrix_id,
        tenant_id="tenant-example",
        owner_user_id="user-jane",
        name="Direct Review",
        knowledge_config_ids=["knowledge-box-matters"],
        document_ids=["doc-box-complaint-outline"],
        columns=[
            ReviewColumn(
                id="deadline",
                label="Deadline",
                question="reviewneedle response deadline",
            )
        ],
        model_id="gpt-4o-mini",
    )


def _fresh_seed_store():
    get_store.cache_clear()
    seed_store = get_store()
    _activate_model_provider(seed_store)
    _add_page_chunks(seed_store)
    return seed_store


def _activate_model_provider(seed_store) -> None:
    provider = seed_store.providers["provider-azure"]
    provider.connected = True
    provider.last_sync = "Loaded for test"
    seed_store.create_provider_key(
        key_id="key-review-test",
        provider=provider,
        name="Review test key",
        environment="Test",
        status="Active",
        expires="Not set",
        secret_value="review-test-secret",
    )


def _add_page_chunks(seed_store) -> None:
    chunks = []
    for page, document_id in enumerate(
        [
            "doc-box-complaint-outline",
            "doc-box-discovery-plan",
            "doc-box-client-update",
        ],
        start=2,
    ):
        chunks.append(
            KnowledgeChunk(
                id=f"review-page-{document_id}",
                knowledge_config_id="knowledge-box-matters",
                document_id=document_id,
                tenant_id="tenant-example",
                source_name=f"{document_id}.pdf",
                source_uri=f"box://review/{document_id}",
                source_type="box",
                text=(
                    "reviewneedle response deadline exposure amount forced failure "
                    "supported contract evidence"
                ),
                page_start=page,
                page_end=page,
                locator=f"Paragraph {page}",
                acl_group_ids=["group-litigation"],
                updated_at="2026-07-20T12:00:00Z",
            )
        )
    seed_store.vector_store.upsert_sources([], chunks)


def _citation(k_index: int) -> ReviewCitation:
    return ReviewCitation(
        k_index=k_index,
        knowledge_config_id="kc-1",
        document_id="doc-1",
        chunk_id=f"chunk-{k_index}",
        source_name="Source.pdf",
        source_uri="https://example.test/Source.pdf",
    )


@pytest.mark.parametrize("answer", ["Grounded [K1].", "Grounded 【K1】.", "Grounded ［K1］."])
def test_citations_are_recognized_in_full_width_brackets(answer: str) -> None:
    # The same model that writes "[K1]" writes "【K1】" in other contexts.
    # Matching only ASCII discarded correct, grounded answers as uncited.
    used = _used_citations(answer, [_citation(1)])

    assert [citation.k_index for citation in used] == [1]


def test_one_invented_index_does_not_discard_an_otherwise_grounded_answer() -> None:
    used = _used_citations("Backed by [K1], and also [K9].", [_citation(1)])

    assert [citation.k_index for citation in used] == [1]


def test_an_answer_citing_only_unknown_indexes_still_fails() -> None:
    with pytest.raises(ReviewCellExecutionError) as raised:
        _used_citations("Backed by [K9].", [_citation(1)])

    assert "[K9]" in str(raised.value)


def test_chat_citations_drop_chunks_the_answer_never_referenced() -> None:
    """Retrieval supplies several chunks per turn; returning all of them made an
    answer look like it cited sources it never used."""
    from app.models.schemas import ChatCitation
    from app.routes.chat import _citations_actually_referenced

    def knowledge(k_index: int) -> ChatCitation:
        return ChatCitation(
            id=f"cite-{k_index}",
            source_name=f"doc-{k_index}.pdf",
            source_type="knowledge",
            source_uri="https://example.test/doc",
            snippet="chunk text",
            k_index=k_index,
        )

    web = ChatCitation(
        id="cite-web-1",
        source_name="Result",
        source_type="web",
        source_uri="https://example.test",
        snippet="snippet",
    )
    supplied = [knowledge(1), knowledge(2), web]

    only_first = _citations_actually_referenced("Grounded in [K1].", supplied)
    assert [c.id for c in only_first] == ["cite-1", "cite-web-1"]

    # The same model emits full-width brackets in other contexts.
    wide = _citations_actually_referenced("Grounded in 【K2】.", supplied)
    assert [c.id for c in wide] == ["cite-2", "cite-web-1"]

    # No marker at all means the model did not use the convention; keeping every
    # citation is better than returning none.
    unmarked = _citations_actually_referenced("A plain answer.", supplied)
    assert [c.id for c in unmarked] == ["cite-1", "cite-2", "cite-web-1"]
